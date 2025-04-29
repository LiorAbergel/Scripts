import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, timedelta
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException
from urllib.parse import urljoin
import cloudscraper
import re
import csv
import os
import subprocess
import platform
from abc import ABC, abstractmethod
from dotenv import load_dotenv
from openpyxl.styles import NamedStyle
from tqdm import tqdm

# Environment Variables -------------------------------------------------------
load_dotenv()
OUTPUT_FILE = os.getenv('OUTPUT_FILE')
CONFIGS_FILE = os.getenv('CONFIGS_FILE')
MIN_DATE = datetime.strptime('05-01-25', "%d-%m-%y").date()
# -----------------------------------------------------------------------------

# Classes ---------------------------------------------------------------------
class Article:
    def __init__(self, title, url, date, tags, source, description=None, image=None, content=None):
        self.title = title
        self.url = url
        self.date = date
        self.source = source
        self.tags = tags
        self.description = description
        self.image = image
        self.content = content

    def __eq__(self, other):
        return isinstance(other, Article) and self.title == other.title and self.url == other.url

    def __hash__(self):
        return hash((self.title, self.url))

    def __repr__(self):
        return f"Article(title='{self.title}', url='{self.url}', date='{self.date}', source='{self.source}')"

class BaseScraper(ABC):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the BaseScraper with the base URL and minimum date."""
        self.base_url = base_url
        self.min_date = min_date
        self.driver = driver or self.create_webdriver()

    def create_webdriver(self):
        """Function to create a headless Chrome WebDriver."""
        chrome_options = Options()
        chrome_options.add_argument("--headless")  # Run in headless mode (no GUI)
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        driver = webdriver.Chrome(options=chrome_options)
        return driver
    
    def fetch_page_content(self, url, use_cloudscraper=False):
        """Fetches the content of a web page using either requests or cloudscraper."""
        if use_cloudscraper:
            scraper = cloudscraper.create_scraper()
            response = scraper.get(url)
        else:
            headers = {'User-Agent': 'Mozilla/5.0'}
            response = requests.get(url, headers=headers)

        if response.status_code == 200:
            return response.text
        else:
            print(f"Failed to retrieve {url}")
            return None

    def parse_html(self, content):
        """Parses the HTML content using BeautifulSoup and returns a soup object."""
        return BeautifulSoup(content, 'html.parser')

    def extract_text(self, tag, selector, default=''):
        """Helper to extract text from a tag using a CSS selector."""
        try:
            element = tag.select_one(selector)
            return element.get_text(strip=True) if element else default
        except Exception as e:
            print(f"Failed to extract text using selector '{selector}': {e}")
            return default
    
    def extract_link(self, tag, selector='', base_url='', default=''):
        """Helper to extract a link from a tag using a CSS selector or directly from the element."""
        # If a selector is provided, find the element using the selector
        if selector:
            tag = tag.select_one(selector)
        
        # Ensure the tag is valid and has the href attribute
        if tag and 'href' in tag.attrs:
            # Use urljoin to handle relative links correctly
            return urljoin(base_url, tag['href'])
        
        return default
    
    def extract_image(self, tag, selector, attribute='src', default=''):
        """Helper to extract an image URL from a tag using a CSS selector."""
        try:
            element = tag.select_one(selector)
            if element:
                if attribute != 'src' and attribute in element.attrs:
                    image = element[attribute]
                elif 'src' in element.attrs:
                    image = element['src']
                elif 'data-cfsrc' in element.attrs:
                    image = element['data-cfsrc']
                elif 'data-srcset' in element.attrs:
                    image = element['data-srcset']
            else:
                image = default

            return image
        except Exception as e:
            print(f"Failed to extract image using selector '{selector}': {e}")
            return default
    
    def extract_content(self, soup, content_selector):
        """Extracts the full content of an article based on the provided CSS selector."""
        content_div = soup.select_one(content_selector)
        if content_div:
            paragraphs = content_div.find_all(['p', 'h2', 'h3', 'li'])
            # Process each paragraph separately
            processed_paragraphs = []
            for p in paragraphs:
                text = p.get_text(strip=True)
                if text:
                    # Replace newlines and commas within paragraphs
                    text = text.replace('\n', ' ').replace(',', ' ')
                    processed_paragraphs.append(text)
            # Join paragraphs with a special delimiter
            text = " || ".join(processed_paragraphs)
            return re.sub(r'[^\x20-\x7E]+', '', text)
        return ""

    def parse_date(self, date_str, date_format='%Y-%m-%d'):
        """Helper to parse a date string into a standardized format."""
        try:
            return datetime.strptime(date_str, date_format).date()
        except ValueError:
            return 'No Date'
    
    def is_article_valid(self, article_date):
        """Validate if the article's date is within the specified minimum date range."""
        return isinstance(article_date, str) or (article_date and article_date >= self.min_date)

    @abstractmethod
    def extract_articles(self, soup):
        """Abstract method to extract articles from the parsed HTML soup."""
        pass

    @abstractmethod
    def scrape(self):
        """Main method to control the scraping process. To be implemented by child classes."""
        pass

class CSOOnlineScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the scraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup, page_number):
        """Extract articles from the soup object."""
        articles = []

        # Scrape articles in the latest-content section (only for the first page)
        if page_number == 1:
            latest_section = soup.find('section', class_='latest-content')
            if latest_section:
                articles.extend(latest_section.find_all('a', class_='card'))
            else:
                latest_section = soup.find('section', class_='featured-content')
                if latest_section:
                    articles.extend(latest_section.find_all('a', class_='featured-content__card'))

        # Scrape articles in the "Articles" section
        articles_section = soup.find('div', class_='content-listing-articles')
        if not articles_section:
            articles_section = soup.find('div', class_='content-listing-various__container')

        if articles_section:
            articles.extend(articles_section.find_all('a', class_='content-row-article'))

        return articles

    def extract_article_content(self, article_url, ):
        """Fetch and extract content and image from the individual article page."""
        # Fetch the article page content
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None, None

        # Parse the HTML content with BeautifulSoup
        soup = self.parse_html(page_content)

        # Extract the full article content using extract_content
        content = self.extract_content(soup, 'div.article__main')

        # Extract the main image using extract_image
        image = self.extract_image(soup, 'div.article__main img', 'src')

        return content, image
    
    def extract_article_details(self, article):
        """Extract details of a single article using helper methods."""
        date_str = self.extract_text(article, 'div.card__info.card__info--light span')
        date = self.parse_date(date_str, date_format="%b %d, %Y")
        if not self.is_article_valid(date):
            return None
        
        title = self.extract_text(article, 'h3.card__title')
        link = self.extract_link(article, '', base_url=self.base_url)
        tags = [tag.get_text(strip=True) for tag in article.find_all('span', class_='card__tag')]
        description = self.extract_text(article, 'p.card__description')

        content, image = self.extract_article_content(link)

        return Article(title, link, date, tags, 'CSOOnline', description, image, content)
    
    def scrape_category(self, url):
        """Scrape a single category from the given URL."""
        all_articles = []
        page_number = 0
        stop_iteration = False

        while not stop_iteration:
            # Increment the page number for subsequent requests
            page_number += 1

            # Construct the URL for the current page
            current_url = f"{url}page/{page_number}" if page_number > 1 else url

            # Fetch page content
            page_content = self.fetch_page_content(current_url)
            if not page_content:
                break

            # Parse the HTML content with BeautifulSoup
            soup = self.parse_html(page_content)
            articles = self.extract_articles(soup, page_number)

            if not articles:
                break
            
            # Process all articles found on the page
            for article in articles:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    stop_iteration = True
                    break

        return all_articles

    def scrape(self):
        """Main method to scrape multiple categories from CSO Online."""
        all_articles = []  # To hold articles from all categories
        categories = [
            "news-analysis", "critical-infrastructure", "cybercrime",
            "network-security", "risk-management", "security-infrastructure",
            "security", "application-security", "vulnerabilities"]
        
        # Generate URLs for each category
        urls = [f"{self.base_url}{category}/" for category in categories]

        # Scrape each category URL
        for category_url in urls:
            category_articles = self.scrape_category(category_url)
            all_articles.extend(category_articles)

        return all_articles

class HackerNewsScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the HackerNewsScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract articles from the soup object."""
        return soup.select('div.body-post.clear')

    def extract_article_content(self, article_url):
        """Fetch and extract content and image from the individual article page."""
        # Fetch the article page content
        page_content = self.fetch_page_content(article_url, use_cloudscraper=True)
        if not page_content:
            return None, None

        # Parse the HTML content with BeautifulSoup
        soup = self.parse_html(page_content)

        content = self.extract_content(soup, '#articlebody')

        image = self.extract_image(soup, 'div.separator img', 'data-src')

        return content, image
    
    def extract_article_details(self, article_tag):
        """Extract the details of a single article."""
        date_str = self.extract_text(article_tag, 'span.h-datetime', default='No Date')
        date_str = date_str.split('')[-1].strip()  # Clean the string
        date = self.parse_date(date_str, date_format="%b %d, %Y")
        if not self.is_article_valid(date):
            return None

        title = self.extract_text(article_tag, 'a.story-link h2.home-title', 'No Title')
        link = self.extract_link(article_tag, 'a.story-link', base_url=self.base_url)
        tags = self.extract_text(article_tag, 'span.h-tags', default='No Tags')

        content, image = self.extract_article_content(link)

        return Article(title, link, date, tags, 'HackerNews', image=image, content=content)

    def get_next_page_url(self, soup, current_url):
        """Find and return the next page URL from the soup object."""
        next_page_tag = soup.select_one('a.blog-pager-older-link-mobile')
        return self.extract_link(next_page_tag, '', base_url=current_url, default=None)

    def scrape(self):
        """Main method to control the scraping process."""
        all_articles = []
        next_page_url = self.base_url
        page_number = 1

        while next_page_url:
            # Fetch the page content
            page_content = self.fetch_page_content(next_page_url, use_cloudscraper=True)
            if not page_content:
                break

            # Parse the HTML content
            soup = self.parse_html(page_content)

            # Extract and process articles from the current page
            articles = self.extract_articles(soup)
            for article_tag in articles:
                article = self.extract_article_details(article_tag)
                if article:
                    all_articles.append(article)
                else:
                    return all_articles  # Stop if an invalid article is encountered

            # Get the next page URL
            next_page_url = self.get_next_page_url(soup, next_page_url)
            page_number += 1

        return all_articles

class IncibeScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the IncibeScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract articles from the parsed HTML soup object."""
        return soup.find_all('article', class_='node vista-blog-resumen node--type-blog node--view-mode-teaser container')

    def extract_article_content(self, article_url, ):
        """Fetch and extract content and image from the individual article page."""
        # Fetch the article page content
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None, None

        # Parse the HTML content with BeautifulSoup
        soup = self.parse_html(page_content)
        article_soup = soup.find("div", class_="node__content")
        
        # Extract the full article content using extract_content
        content = self.extract_content(article_soup, 'div.clearfix.text-formatted.field.field--name-body.field--type-text-with-summary.field--label-hidden.field__item')

        # Extract the main image using extract_image
        image = "www.incibe.es" + self.extract_image(soup, 'div.field.field--name-field-noticia-fotografia img', 'src')

        return content, image
    
    def extract_article_details(self, article):
        """Extract the details of a single article."""
        # Extract and clean the date
        date_str = self.extract_text(article, 'div.node__content.postedOnLabel', default='').split('by')[0].strip()
        date_str = date_str.replace('Updated on', '').replace('Posted on', '').replace(',', '').strip()
        date = self.parse_date(date_str, date_format="%d/%m/%Y")
        if not self.is_article_valid(date):
            return None

        title = self.extract_text(article, 'h3 a')
        link = self.extract_link(article, 'h3 a', base_url=self.base_url)
        tags = [tag.get_text(strip=True) for tag in article.select('div.field--name-field-tax-etiquetas li.tag_clouds_term a')]
        description = self.extract_text(article, 'div.clearfix.text-formatted.field.field--name-body.field--type-text-with-summary.field--label-hidden.field__item')

        content, image = self.extract_article_content(link)

        return Article(title, link, date, tags, 'Incibe', description, image, content)

    def get_next_page_url(self, page_count):
        """Construct the next page URL."""
        return f"{self.base_url}?page={page_count}"

    def scrape(self):
        """Main method to control the scraping process."""
        all_articles = []
        next_page_url = self.base_url
        page_count = 0

        while next_page_url:
            # Fetch the page content
            page_content = self.fetch_page_content(next_page_url)
            if not page_content:
                break

            # Parse the HTML content
            soup = self.parse_html(page_content)

            # Extract and process articles from the current page
            articles = self.extract_articles(soup)
            for article in articles:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    return all_articles  # Stop if an invalid article is encountered

            # Get the next page URL
            page_count += 1
            next_page_url = self.get_next_page_url(page_count)

        return all_articles

class SecurityWeekScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the SecurityWeekScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)
        self.categories = ["ics-ot", "iot-security"]

    def calculate_article_date(self, relative_time_str):
        """Function to calculate the article date based on the relative time string."""
        # Get today's date
        today = datetime.now()

        # Regular expressions to extract the time and unit from the string
        time_pattern = r'(\d+)\s*(day|hour|minute|second|week|month|year)s?\s*ago'

        # Search for the pattern in the input string
        match = re.search(time_pattern, relative_time_str)
        
        if not match:
            return None  # If no match, return None or handle it differently

        # Extract the quantity and the time unit
        quantity = int(match.group(1))
        time_unit = match.group(2)

        # Calculate the date based on the time unit
        if time_unit == 'day':
            article_date = today - timedelta(days=quantity)
        elif time_unit == 'hour':
            article_date = today - timedelta(hours=quantity)
        elif time_unit == 'minute':
            article_date = today - timedelta(minutes=quantity)
        elif time_unit == 'second':
            article_date = today - timedelta(seconds=quantity)
        elif time_unit == 'week':
            article_date = today - timedelta(weeks=quantity)
        elif time_unit == 'month':
            article_date = today - timedelta(days=30 * quantity)  # Approximation: 1 month = 30 days
        elif time_unit == 'year':
            article_date = today - timedelta(days=365 * quantity)  # Approximation: 1 year = 365 days
        else:
            return None  # Return None if an unknown unit is encountered

        return article_date.date()

    def extract_articles(self, soup):
        """Extract articles from the parsed HTML soup."""
        # Combine the articles from both upper and lower sections
        upper_section = soup.select_one('div.zox-feat-tech2-grid.left.zoxrel.zox100')
        lower_section = soup.select_one('div.zox-main-blog.zoxrel.left.zox100')

        upper_articles = upper_section.find_all('section',class_='zox-art-wrap') if upper_section else []
        lower_articles = lower_section.find_all('article', class_='zox-art-wrap') if lower_section else []

        return upper_articles + lower_articles

    def extract_article_content(self, article_url):
        """Fetch and extract content and image from the individual article page."""
        # Fetch the article page content
        page_content = self.fetch_page_content(article_url, use_cloudscraper=True)
        if not page_content:
            return None, None

        soup = BeautifulSoup(page_content, 'html.parser')

        # Extract the description from the excerpt
        description = self.extract_text(soup, 'span.zox-post-excerpt')

        # Extract the content paragraphs
        content = self.extract_content(soup, 'div.zox-post-body')

        # Extract the image URL
        image_url = self.extract_image(soup, 'div.zox-post-img-wrap img', 'src')

        return description, content, image_url

    def extract_article_details(self, article):
        """Extract details for a single article."""
        # Extract title and link
        title = self.extract_text(article, 'h2.zox-s-title2', default='No title')
        link = self.extract_link(article, 'a', base_url=self.base_url, default='No link')

        # Extract the article date
        date_text = self.extract_text(article, 'span.zox-byline-date', default='No date')
        try:
            article_date = self.calculate_article_date(date_text)
            if not article_date:
                article_date = datetime.strptime(date_text, "%B %d, %Y").date()
        except ValueError:
            print(f"Error parsing date: {date_text}")
            return None

        # Validate the article's date
        if not self.is_article_valid(article_date):
            return None

        # Extract tags
        category = self.extract_text(article, 'h3.zox-s-cat', default='')

        description, content, image = self.extract_article_content(link)

        return Article(title, link, article_date, category, 'SecurityWeek', description, image, content)

    def scrape_category(self, category_url):
        """Scrape articles from a single category page."""
        all_articles = []
        page_count = 1
        stop_scraping = False

        while not stop_scraping:
            # Construct the URL for the current page
            current_url = f"{category_url}page/{page_count}" if page_count > 1 else category_url

            # Fetch and parse the page content
            page_content = self.fetch_page_content(current_url, use_cloudscraper=True)
            if not page_content:
                break

            soup = self.parse_html(page_content)
            articles = self.extract_articles(soup)

            # If no articles are found, stop scraping this category
            if not articles:
                break

            # Extract details for each article
            for article in articles:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    stop_scraping = True
                    break

            page_count += 1

        return all_articles

    def scrape(self):
        """Main method to scrape multiple categories from SecurityWeek."""
        all_articles = []

        # Scrape each category
        for category in self.categories:
            category_url = f"{self.base_url}{category}/"
            category_articles = self.scrape_category(category_url)
            all_articles.extend(category_articles)

        return all_articles

class DragosScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the DragosScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract articles from the parsed HTML soup object."""
        # Find all article containers in the main content area
        article_divs = soup.find('div', class_='facetwp-template row facetwp_blog_posts')
        return article_divs.find_all('article', class_='blog-posts-article') if article_divs else []

    def extract_article_content(self, article_url):
        """Fetch and extract content and image from the individual article page."""
        # Fetch the article page content
        self.driver.get(article_url)
        time.sleep(3)  # Wait for the page to load

        # Parse the page source using BeautifulSoup
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')

        # Extract the main article content using the extract_text helper method
        content = self.extract_content(soup, 'div.col-md-8.single-post__main.px-0')

        return content

    def extract_article_details(self, article):
        """Extract and return the details of a single article using helper functions from BaseScraper."""
        # Extract the article date using helper functions
        date_str = self.extract_text(article, 'p.meta.meta--bottom', default='').split('|')[0].strip()
        article_date = self.parse_date(date_str, "%m.%d.%y")

        if not self.is_article_valid(article_date):
            return None

        # Extract the article title and URL using helper functions
        title = self.extract_text(article, 'h5.blog-posts-article--title a')
        article_url = self.extract_link(article, 'h5.blog-posts-article--title a', base_url=self.base_url)
        tags = self.extract_text(article, 'div.blog-posts-topic--container', default='No tags')

        image_tag = article.find('div', class_='blog-posts-thumbnail')['style']
        image = re.search(r'url\((.*?)\)', image_tag).group(1).strip() if image_tag else ''

        content = self.extract_article_content(article_url)

        # Create and return the Article object
        return Article(title, article_url, article_date, tags, 'Dragos', image=image, content=content)

    def get_next_page_button(self):
        """Check and return the 'Load More' button if available."""
        try:
            load_more_button = self.driver.find_element(By.CLASS_NAME, 'facetwp-load-more')
            return load_more_button if load_more_button.is_displayed() else None
        except Exception as e:
            print(f"Error finding 'Load More' button: {e}")
            return None

    def scrape_page(self, initial_article_count):
        """Scrape a single page of articles."""
        # Parse the page source using BeautifulSoup
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        articles = self.extract_articles(soup)

        # Process and extract details from each article
        article_list = []
        for article in articles[initial_article_count:]:
            article_details = self.extract_article_details(article)
            if article_details:
                article_list.append(article_details)

        return article_list

    def scrape(self):
        """Main method to control the scraping process."""
        all_articles = []
        load_more_count = 1
        self.driver.get(self.base_url)
        time.sleep(3)  # Let the page load

        # Continuously click "Load More" until no more articles are available
        while True:
            initial_article_count = len(all_articles)

            # Scrape articles from the current page
            articles = self.scrape_page(initial_article_count)
            if articles:
                all_articles.extend(articles)
            
            # If no new articles are added, break the loop (prevents infinite loops)
            if len(all_articles) == initial_article_count:
                break

            # Check and click the "Load More" button if available
            self.driver.get(self.base_url) # Load the base URL
            for i in range(load_more_count):
                load_more_button = self.get_next_page_button()
                if load_more_button:
                        self.driver.execute_script(f"window.scrollBy(0, {-100 * (i + 1)});") 
                        load_more_button.click()
                        time.sleep(3)  # Wait for the new articles to load
                else:
                    break  # No more "Load More" button, stop the loop
            load_more_count += 1

        return all_articles

class TrendMicroScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the TrendMicroScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)
        self.load_more_count = 1  # Track how many times 'Load More' is clicked

    def handle_cookies(self):
        """Function to handle cookies or pop-ups on the page."""
        try:
            cookie_button = self.driver.find_element(By.ID, 'onetrust-accept-btn-handler')
            if cookie_button.is_displayed():
                cookie_button.click()
                time.sleep(2)  # Wait for the banner to disappear
        except Exception as e:
            print(f"No cookies/banner found or error encountered: {e}")
    
    def get_next_page_button(self):
        """Check and return the 'Load More' button if available."""
        try:
            load_more_button = self.driver.find_element(By.CLASS_NAME, 'load-more-btn')
            return load_more_button if load_more_button.is_displayed() else None
        except Exception as e:
            print(f"Error finding 'Load More' button: {e}")
            return None

    def extract_article_content(self, article_url):
        """Extract and return the main content of the article."""
        self.driver.get(article_url)
        soup = self.parse_html(self.driver.page_source)
        return self.extract_content(soup, 'main.main--content')

    def extract_articles(self, soup):
        """Extract articles from the parsed HTML soup object."""
        return soup.find_all('article', class_='grid-item')
    
    def extract_featured_article(self, soup):
        """Extract the featured article from the page."""
        featured_article_tag = soup.find('div', class_='featured-article')
        if not featured_article_tag:
            return None
        
        date_text = self.extract_text(featured_article_tag, 'div.properties span', default='No date')
        article_date = self.parse_date(date_text, date_format="%B %d, %Y")
        if not self.is_article_valid(article_date):
            return None

        # Use helper methods from BaseScraper to extract details
        title = self.extract_text(featured_article_tag, 'h2.article-title a', default='No title')
        article_url = self.extract_link(featured_article_tag, 'h2.article-title a', base_url=self.base_url, default='No link')
        tags = self.extract_text(featured_article_tag, 'div.article-filter-tag', default='No tags')
        description = self.extract_text(featured_article_tag, 'p.article-description', default='No description')
        image_url = self.extract_image(featured_article_tag, 'figure.img-container img', default='No image')
        content = self.extract_article_content(article_url)

        return Article(title, article_url, article_date, tags, 'TrendMicro', description, image_url, content)

    def extract_article_details(self, tag):
        """Extract and return the details of a single article using helper methods from BaseScraper."""
        # Extract and parse the publication date
        date_spans = tag.find('div', class_='properties').find_all('span')
        index = -1 if len(date_spans) > 2 else 0
        date_text = date_spans[index].get_text(strip=True)
        article_date = self.parse_date(date_text, date_format="%b %d, %Y")
        if not self.is_article_valid(article_date):
            return None
        
        title = self.extract_text(tag, 'h3.heading a', default='No title')
        article_url = self.extract_link(tag, 'h3.heading a', base_url=self.base_url, default='No link')
        tags = self.extract_text(tag, 'div.filter-tag', default='No tags')
        description = self.extract_text(tag, 'p.description', default='No description')
        image_url = self.extract_image(tag, 'figure.img-container img', default='No image')
        content = self.extract_article_content(article_url)

        return Article(title, article_url, article_date, tags, 'TrendMicro', description, image_url, content)

    def scrape_page(self, initial_article_count):
        """Scrape articles from the current page after the specified initial article count."""
        soup = self.parse_html(self.driver.page_source)
        article_list = []

        if initial_article_count == 0:
            featured_article = self.extract_featured_article(soup)
            if featured_article:
                article_list.append(featured_article)
        
        articles = self.extract_articles(soup)

        # Only process articles that come after the initial count
        for article in articles[initial_article_count:]:
            article_details = self.extract_article_details(article)
            if article_details:
                article_list.append(article_details)

        return article_list

    def load_articles_until_min_date(self):
        """Load articles by clicking 'Load More' until articles older than the minimum date are found."""
        all_articles = []
        self.handle_cookies()  # Handle cookies before scraping
        load_more_count = 1
        
        while True:
            initial_article_count = len(all_articles)  # Track number of already scraped articles

            # Scrape articles on the current page
            articles = self.scrape_page(initial_article_count)
            if articles:
                all_articles.extend(articles)

            # If no new articles were added, break the loop to prevent infinite loading
            if len(all_articles) == initial_article_count:
                break

            # Check and click the "Load More" button if available
            self.driver.get(self.base_url) # Load the base URL
            for _ in range(load_more_count):
                load_more_button = self.get_next_page_button()
                if load_more_button:
                        load_more_button.click()
                        time.sleep(3)  # Wait for the new articles to load
                else:
                    break  # No more "Load More" button, stop the loop
            load_more_count += 1

        return all_articles

    def scrape(self):
        """Main method to control the scraping process."""
        all_articles = []
        try:
            self.driver.get(self.base_url)
            time.sleep(3)  # Allow the page to fully load

            # Load articles until an article older than the minimum date is found
            all_articles = self.load_articles_until_min_date()
        except Exception as e:
            print(f"Error during scraping process: {e}")

        return all_articles

class CISAScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the CISAScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract articles from the parsed HTML soup object."""
        # Locate all article containers on the page
        return soup.find_all('article', class_='is-promoted c-teaser c-teaser--horizontal')

    def extract_article_details(self, container):
        """Extract and return the details of a single article from a container."""
        date_text = self.extract_text(container, 'div.c-teaser__date')
        article_date = self.parse_date(date_text, "%b %d, %Y")
        if not self.is_article_valid(article_date):
            return None

        title = self.extract_text(container, 'h3.c-teaser__title')
        article_url = self.extract_link(container, 'h3.c-teaser__title a', base_url="https://www.cisa.gov")
        tags = self.extract_text(container, 'div.c-teaser__meta').split("|")

        content = self.extract_article_content(article_url)

        return Article(title, article_url, article_date, tags, 'CISA', content=content)

    def extract_article_content(self, article_url):
        """Fetch and extract the main content of the article."""
        # Fetch the article page content
        page_content = self.fetch_page_content(article_url, use_cloudscraper=True)
        if not page_content:
            return None

        # Parse the HTML content using BeautifulSoup
        soup = self.parse_html(page_content)

        return self.extract_content(soup, 'div.l-page-section__content')

    def scrape_page(self, url):
        """Scrape a single page and return articles from that page."""
        articles = []
        page_content = self.fetch_page_content(url, use_cloudscraper=True)

        if not page_content:
            return articles

        # Parse the HTML content with BeautifulSoup
        soup = self.parse_html(page_content)

        # Find and process all article containers on the page
        article_containers = self.extract_articles(soup)
        if not article_containers:
            return articles

        # Extract article details for each article container
        for container in article_containers:
            article_details = self.extract_article_details(container)
            if article_details:
                articles.append(article_details)

        return articles

    def scrape(self):
        """Main method to control the scraping process for the entire CISA website."""
        all_articles = []
        current_page = 1

        while True:
            # Construct the URL for the current page
            url = f"{self.base_url}?page={current_page}" if current_page > 1 else self.base_url

            # Scrape the current page for articles
            page_articles = self.scrape_page(url)

            # If no articles found or if all are older than min_date, stop scraping
            if not page_articles:
                break

            all_articles.extend(page_articles)
            current_page += 1

        return all_articles

class TheRecordScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the scraper with the base URL, minimum date, and Selenium WebDriver."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract articles from the parsed HTML soup object."""
        return soup.find_all('a', class_='article-tile')

    def extract_article_content(self, article_url):
        """Extract and return the main content of the article."""
        self.driver.get(article_url)

        try:
            # Wait until the main content is fully loaded
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'span.wysiwyg-parsed-content'))
            )
        except TimeoutException:
            print("Main content took too long to load or is missing.")
            return "No content available"

        # Parse the fully loaded page
        soup = self.parse_html(self.driver.page_source).find('div', class_='article__content')

        # Now extract content using the BaseScraper method
        return self.extract_content(soup, 'span.wysiwyg-parsed-content')
    
    def extract_article_details(self, container):
        """Extract and return the details of a single article."""
        # Extract and parse the publication date using BaseScraper's extract_text and parse_date
        date_text = self.extract_text(container, 'span.article-tile__meta__date', default=None)
        try:
            # Clean up date string by removing suffixes like "st", "nd", "rd", "th"
            cleaned_date_text = re.sub(r'(\d+)(st|nd|rd|th)', r'\1', date_text)
            article_date = self.parse_date(cleaned_date_text, "%B %d, %Y")
        except ValueError:
            print(f"Error parsing date: {date_text}")
            return None
        
        # Skip articles published before the minimum date
        if not article_date or not self.is_article_valid(article_date):
            return None
        
        # Extract the title, cleaning it up using BaseScraper's extract_text method
        title = self.extract_text(container, 'h2.article-tile__title', default='No Title')
        if title.startswith("Brief"):
            title = title.replace("Brief", "")

        # Extract the article URL using BaseScraper's extract_link method
        article_url = self.extract_link(container, base_url="https://therecord.media")

        image = self.extract_image(container, 'div.article-tile__img img', 'src')

        content = self.extract_article_content(article_url)

        return Article(title, article_url, article_date, None, 'TheRecord', image=image, content=content)

    def navigate_and_filter(self):
        """Navigate to the site and filter by the 'Cybercrime' category."""
        self.driver.get(self.base_url)
        time.sleep(3)  # Wait for the page to load

        # Handle cookie consent if present
        self.handle_cookie_consent()

        # Handle modal pop-up if present
        self.handle_modal_popup()

    def handle_cookie_consent(self):
        """Handle cookie consent pop-up if present."""
        try:
            cookie_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, "hs-eu-confirmation-button"))
            )
            cookie_button.click()
            time.sleep(2)
        except Exception as e:
            pass

    def handle_modal_popup(self):
        """Handle modal pop-up (e.g., newsletter signup) if present."""
        try:
            # Wait for the modal close button and close the modal if present
            modal_close_button = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.CLASS_NAME, "leadinModal-close"))
            )
            modal_close_button.click()
        except TimeoutException:
            pass

    def get_next_page(self):
        """Click the 'Next' button to navigate to the next page if available, handling any overlay interruptions."""
        try:

            # Locate the next page button
            next_page_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'pagination__btn-next'))
            )

            # Scroll the button into view
            # self.driver.execute_script("arguments[0].scrollIntoView();", next_page_button)

            # Handle any modal pop-ups that may interfere with clicking
            # self.handle_modal_popup()
            
            # Attempt to click, handling potential overlays
            retries = 3
            while retries > 0:
                try:
                    next_page_button.click()
                    time.sleep(2)  # Wait for the page to load
                    return True
                except ElementClickInterceptedException:
                    print("Click intercepted, handling overlay and retrying.")
                    self.handle_modal_popup()  # Attempt to close any overlay
                    retries -= 1
                    time.sleep(1)

            # Return None if unable to click after retries
            print("Failed to click 'Next' button after multiple attempts.")
            return None
        except Exception as e:
            print(f"Error clicking the next page button: {e}")
            return None

    def scrape(self):
        """Main method to control the scraping process."""
        all_articles = []  # Store all articles from all pages
        page_count = 1
        stop_scraping = False

        self.navigate_and_filter()  # Navigate and filter articles by 'Cybercrime'

        while not stop_scraping:
            # Parse the current page content with BeautifulSoup
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            article_containers = self.extract_articles(soup)

            # Extract details for each article
            for container in article_containers:
                article = self.extract_article_details(container)
                if article:
                    all_articles.append(article)
                else:
                    stop_scraping = True  # Stop if an invalid article is encountered

            # Navigate to the next page, stop if no more pages
            self.navigate_and_filter()  # Reapply the 'Cybercrime' filter
            for _ in range(page_count):
                next_page_button = self.get_next_page()
                if not next_page_button:
                    stop_scraping = True
                    break
            
            page_count += 1

        return all_articles

class BleepingComputerScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the BleepingComputerScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)
        self.page_number = 1  # Initialize page number for pagination
    
    def extract_article_content(self, article_url):
        """Fetch and extract content and image from the individual article page."""
        # Fetch the article page content
        page_content = self.fetch_page_content(article_url, True)
        if not page_content:
            return None, None

        # Parse the HTML content with BeautifulSoup
        soup = self.parse_html(page_content)

        # Extract the main article content using extract_content
        content = self.extract_content(soup, 'div.articleBody')

        # Extract the main image using extract_image
        image = self.extract_image(soup, 'div.articleBody p img', 'src')

        return content, image

    def extract_articles(self, soup):
        """Extract articles from the parsed HTML soup object."""
        # Locate the main article container using the ID "bc-home-news-main-wrap"
        main_articles_container = soup.find('ul', id='bc-home-news-main-wrap')
        if not main_articles_container:
            return []

        # Return all article containers found in the main section
        return main_articles_container.find_all('div', class_='bc_latest_news_text')

    def extract_article_details(self, article):
        """Extract details from an article element and return an Article object."""
        date_str = self.extract_text(article, 'li.bc_news_date', default='No Date')
        article_date = self.parse_date(date_str, date_format="%B %d, %Y")
        if not self.is_article_valid(article_date):
            return None
        
        title = self.extract_text(article, 'h4 a', default='No Title')
        link = self.extract_link(article, 'h4 a', base_url=self.base_url)
        tags = [self.extract_text(article, 'div.bc_latest_news_category span a', default='No Tag')]
        if 'Deals' in tags:
            return "pass"  # Skip articles tagged as "Deals"
        description = self.extract_text(article, 'p', default='No description available')

        content, image = self.extract_article_content(link)

        return Article(title, link, article_date, tags, 'BleepingComputer', description=description, image=image, content=content)

    def scrape(self):
        """Main method to scrape articles with pagination."""
        all_articles = []
        stop_scraping = False
        
        while not stop_scraping:
            current_url = self.build_page_url()
            page_content = self.fetch_page_content(current_url, True)
            if not page_content:
                print(f"Failed to retrieve content from {current_url}. Stopping.")
                break

            soup = self.parse_html(page_content)
            article_items = self.extract_articles(soup)
            if not article_items:
                print(f"No articles found on page {self.page_number}. Stopping.")
                break

            # Process each article
            for article in article_items:
                time.sleep(2)
                article_details = self.extract_article_details(article)
                if article_details == "pass":
                    continue
                if article_details:
                    all_articles.append(article_details)
                else:
                    stop_scraping = True
                    break

            # Move to the next page
            self.page_number += 1

        return all_articles

    def build_page_url(self):
        """Construct the URL for the current page."""
        return f"{self.base_url}page/{self.page_number}/" if self.page_number > 1 else self.base_url

class InfoSecurityMagazineScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the InfoSecurityMagazineScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)
    
    def extract_article_content(self, article_url):
        """Fetch and extract content from the individual article page."""
        # Fetch the article page content
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None

        # Parse the HTML content with BeautifulSoup
        soup = self.parse_html(page_content)

        # Extract the main article content from paragraphs within the main content div
        content = self.extract_content(soup, 'div.content-module')

        return content

    def extract_articles(self, soup):
        """Extract articles from the main article container in the soup object."""
        main_articles_container = soup.find('ol', class_='webpages-list')
        if not main_articles_container:
            return []  # Return an empty list if the container is not found
        return main_articles_container.find_all('li', class_='webpage-item')

    def extract_article_details(self, article):
        """Extract the details of a single article and return an Article object."""
        # Extract the article date using the BaseScraper date parsing method
        date_str = self.extract_text(article, 'time', default=None)
        article_date = self.parse_date(date_str, date_format="%d %b %Y")
        if not self.is_article_valid(article_date):
            return None 
        
        # Extract the title and link using the BaseScraper helper methods
        title = self.extract_text(article, 'h2.h3.webpage-title', default='No Title')
        link = self.extract_link(article, 'h2.h3.webpage-title a', base_url=self.base_url)

        description = self.extract_text(article, 'p.webpage-summary', default='No description available')

        image = self.extract_image(article, 'img.webpage-thumb', 'src')

        content = self.extract_article_content(link)

        # Return the Article object
        return Article(title, link, article_date, None, 'InfoSecurity Magazine', description, image, content)

    def scrape(self):
        """Main method to scrape articles from InfoSecurity Magazine with pagination support."""
        all_articles = []  # To store all articles
        page_number = 1
        stop_scraping = False

        while not stop_scraping:
            # Construct the URL for the current page
            current_url = f"{self.base_url}/page-{page_number}/" if page_number > 1 else self.base_url

            # Fetch page content
            page_content = self.fetch_page_content(current_url)
            if not page_content:
                print(f"Failed to retrieve content from {current_url}. Stopping.")
                break

            # Parse the HTML content with BeautifulSoup
            soup = self.parse_html(page_content)

            # Extract articles from the current page
            article_items = self.extract_articles(soup)
            if not article_items:
                print(f"No articles found on page {page_number}. Stopping.")
                break

            # Process each article and stop if an old article is found
            for article in article_items:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    stop_scraping = True
                    break

            # Move to the next page
            page_number += 1
            time.sleep(1)  # Add a delay to avoid overloading the server

        return all_articles

class SecurityAffairsScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the SecurityAffairsScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract articles from the main article container in the soup object."""
        main_articles_container = soup.find('div', class_='latest-news-block')
        if not main_articles_container:
            return []  # Return an empty list if the container is not found
        return main_articles_container.find_all('div', class_='news-card news-card-category mb-3 mb-lg-5')

    def extract_article_details(self, article):
        """Extract the details of a single article and return an Article object."""
        # Extract the article date
        date_container = article.find('div', class_='post-time mb-3')
        article_date = self.extract_article_date(date_container)
        if not self.is_article_valid(article_date):
            return None
        
        # Extract the title and link using the BaseScraper helper methods
        title = self.extract_text(article, 'h5.mb-3', default='No Title')
        link = self.extract_link(article, 'h5.mb-3 a', base_url=self.base_url)

        image = self.extract_image(article, 'div.news-card-pic a img', 'src')

        content = self.extract_article_content(link)

        # Return the Article object
        return Article(title, link, article_date, None, 'Security Affairs', image=image,content=content)
    
    def extract_article_content(self, article_url):
        """Fetch and extract content from the individual article page."""
        # Fetch the article page content
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None

        # Parse the HTML content with BeautifulSoup
        soup = self.parse_html(page_content)

        # Extract the main article content from paragraphs within the main content div
        content = self.extract_content(soup, 'div.article-details-block')

        return content 

    def extract_article_date(self, date_container):
        """Helper method to extract and parse the article date."""
        if date_container:
            date_tags = date_container.find_all('span')
            if len(date_tags) > 1:
                date_str = date_tags[1].get_text(strip=True)
                return self.parse_date(date_str, date_format="%B %d, %Y")

    def scrape(self):
        """Main method to scrape articles from Security Affairs with pagination support."""
        all_articles = []
        page_number = 1
        stop_scraping = False

        while not stop_scraping:
            # Construct the URL for the current page
            current_url = f"{self.base_url}/page/{page_number}/" if page_number > 1 else self.base_url

            # Fetch page content
            page_content = self.fetch_page_content(current_url)
            if not page_content:
                print(f"Failed to retrieve content from {current_url}. Stopping.")
                break

            # Parse the HTML content with BeautifulSoup
            soup = self.parse_html(page_content)

            # Extract articles from the current page
            article_items = self.extract_articles(soup)
            if not article_items:
                print(f"No articles found on page {page_number}. Stopping.")
                break

            # Process each article and stop if an old article is found
            for article in article_items:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    stop_scraping = True
                    break

            # Move to the next page
            page_number += 1
            time.sleep(1)  # Add a delay to avoid overloading the server

        return all_articles

class TripWireScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the TripWireScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract articles from the main article container in the soup object."""
        return soup.find_all('article', class_='node node--type-blog node--view-mode-search-index')
    
    def extract_article_content(self, article_url):
        """Fetch and extract content from the individual article page."""
        # Fetch the article page content
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None

        # Parse the HTML content with BeautifulSoup
        soup = self.parse_html(page_content)

        # Extract the main article content from paragraphs within the main content div
        content = self.extract_content(soup, 'div.field.field--name-body')

        return content 

    def extract_article_details(self, article):
        """Extract and return the details of a single article."""
        # Extract the article date using a regular expression
        date_container = article.find('div', class_='node--submitted')
        article_date = self.extract_article_date(date_container)
        if not self.is_article_valid(article_date):
            return None
        
        # Extract the title and link using BaseScraper helper methods
        title = self.extract_text(article, 'h3.node--title', default='No Title')
        link = self.extract_link(article, 'a.order-2.order-md-0', base_url=self.base_url)

        tags = self.extract_tags(article, 'div.field--name-field-cornerstone-relationship div.field__item')

        image_url = self.extract_image(article, 'div.field--name-field-media a img', 'src')

        # Since the image URL is relative, join it with the base URL
        full_image_url = urljoin(self.base_url, image_url)

        content = self.extract_article_content(link)

        # Return the Article object
        return Article(title, link, article_date, tags, 'Tripwire', image=full_image_url, content=content)

    def extract_article_date(self, date_container):
        """Helper method to extract and parse the article date."""
        if date_container:
            # Use regular expression to extract date from text (e.g., "09/30/2024")
            date_match = re.search(r'\d{2}/\d{2}/\d{4}', date_container.get_text(strip=True)).group()
            return self.parse_date(date_match, date_format="%m/%d/%Y")
        return 'No Date'

    def extract_tags(self, article, selector):
        """Extract tags from the article using a CSS selector."""
        tags_container = article.select(selector)
        return [tag.get_text(strip=True) for tag in tags_container] if tags_container else []

    def scrape(self):
        """Main method to scrape articles from Tripwire with pagination support."""
        all_articles = []  # To store all articles
        page_index = 0  # Start from page 0
        stop_scraping = False

        while not stop_scraping:
            # Construct the URL for the current page
            current_url = f"{self.base_url}?page={page_index}"

            # Fetch page content
            page_content = self.fetch_page_content(current_url)
            if not page_content:
                print(f"Failed to retrieve content from {current_url}. Stopping.")
                break

            # Parse the HTML content with BeautifulSoup
            soup = self.parse_html(page_content)

            # Extract articles from the current page
            article_items = self.extract_articles(soup)
            if not article_items:
                print(f"No articles found on page {page_index + 1}. Stopping.")
                break

            # Process each article and stop if an old article is found
            for article in article_items:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    stop_scraping = True
                    break

            # Move to the next page
            page_index += 1
            time.sleep(1)  # Add a delay to avoid overloading the server

        return all_articles

class RockwellAutomationScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the RockwellAutomationScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)
        self.pages_to_scrape = ["the-journal.html", "blogs.html"]

    def extract_articles(self, soup):
        """Extract all article containers from the page."""
        return soup.find_all('div', class_='generic-filter__result')

    def extract_article_content(self, article_url):
        """Extract and return the main content of the article."""
        self.driver.get(article_url)
        soup = self.parse_html(self.driver.page_source)
        return self.extract_content(soup, 'div.generic-container__inner')

    def extract_article_details(self, article, base_url):
        """Extract and return the details of a single article."""
        date_tag = article.find('div', class_='generic-filter__result-date')
        article_date= self.parse_date(date_tag.get_text(strip=True), "%B %d, %Y")
        if not self.is_article_valid(article_date):
            return None
        
        title_tag = article.find('a', class_='generic-filter__result-title')

        title = title_tag.get_text(strip=True)
        link = urljoin(base_url, title_tag['href'])

        description = self.extract_description(article)

        image = self.extract_image(article, 'div.generic-filter__result-image-wrapper img')

        content = self.extract_article_content(link)

        return Article(title, link, article_date, None, 'Rockwell Automation', description, image=image, content=content)

    def extract_description(self, article):
        """Extract description text for the article."""
        description_container = article.find('div', class_='generic-filter__result-description')
        return description_container.get_text(strip=True) if description_container else ""

    def scrape_page(self, driver, base_url):
        """Scrape articles from a single page using the Selenium driver."""
        all_articles = []
        stop_scraping = False

        while not stop_scraping:
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            article_containers = self.extract_articles(soup)

            if not article_containers:
                print(f"No more articles found on {base_url}. Stopping scrape for this page.")
                break

            for container in article_containers:
                article_details = self.extract_article_details(container, base_url)
                if article_details:
                    all_articles.append(article_details)
                else:
                    stop_scraping = True
                    break

            if not self.load_more_articles(driver):
                break

        return all_articles

    def load_more_articles(self, driver):
        """Click the 'Load More' button to load more articles, if available."""
        try:
            load_more_button = driver.find_element(By.CLASS_NAME, 'generic-filter__btn-load-more')
            if load_more_button and load_more_button.is_displayed():
                load_more_button.click()
                time.sleep(3)  # Wait for new articles to load
                return True
            else:
                print("No more articles to load.")
                return False
        except Exception as e:
            print(f"Error clicking the 'Load More' button: {e}")
            return False

    def scrape(self):
        """Main method to scrape articles from multiple pages on Rockwell Automation's site."""
        all_articles = []

        for page in self.pages_to_scrape:
            url = f"{self.base_url}{page}"
            self.driver.get(url)
            time.sleep(3)  # Allow the page to load
            articles = self.scrape_page(self.driver, self.base_url)
            all_articles.extend(articles)

        return all_articles

class NozomiScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the NozomiScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract all article containers from the main resource container."""
        article_container = soup.find('div', class_='resource-collection-list-responsive w-dyn-items')
        if not article_container:
            print("No articles found on the page.")
            return []
        return article_container.find_all('div', role='listitem')

    def extract_article_content(self, article_url):
        """Fetch and extract content and image from the individual article page."""
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None, None, None

        soup = self.parse_html(page_content)

        date_str = self.extract_text(soup, 'div.post-author-text-byline.post-author-text-byline-date') 
        article_date = self.parse_date(date_str, '%B %d, %Y')

        # Extract the main content using extract_content helper
        content = self.extract_content(soup, 'div.post-body-rich-text-block.blog-tich-text.w-richtext')

        # Extract the main image
        image = self.extract_image(soup, 'div.ost-template-body-content-column.w-col.w-col-9.w-col-stack img', 'src')

        return article_date, content, image

    def extract_article_details(self, article):
        """Extract and return the details of a single article."""
        title = self.extract_text(article, 'h2.resource-page-resource-block-item-header')
        link = self.extract_link(article, 'a.blog-page-link-block', base_url=self.base_url)

        # Extract content and image
        article_date, content, image = self.extract_article_content(link)
        if not self.is_article_valid(article_date):
            return None

        return Article(title, link, article_date, None, 'Nozomi', content=content, image=image)

    def scrape(self):
        """Main method to control the scraping process."""
        all_articles = []
        page_content = self.fetch_page_content(self.base_url)

        if not page_content:
            return all_articles

        soup = self.parse_html(page_content)

        # Extract articles from the page
        articles = self.extract_articles(soup)

        # Process and extract details from each article
        for article in articles:
            article_details = self.extract_article_details(article)
            if article_details:
                all_articles.append(article_details)
            else:
                break

        return all_articles

class CyberSecurityDiveScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the CyberSecurityDiveScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract all article containers from the main feed container."""
        return soup.find_all('li', class_='row feed__item')
    
    def extract_featured_articles(self, soup):
        """Extract featured articles from the current page soup object."""
        # Find main featured article section and top stories section
        hero_article = soup.select_one("section.hero-article")
        top_stories = soup.select("section.top-stories ol li")
        
        featured_articles = []

        # Process the main hero article if it exists
        if hero_article:
            featured_articles.append(self.extract_featured_article_details(hero_article, main=True))
        
        # Process each top story if available
        for story in top_stories:
            featured_articles.append(self.extract_featured_article_details(story, main=False))

        return [article for article in featured_articles if article]
    
    def extract_featured_article_details(self, article, main=True):
        """Extract details for a featured article. Different layout handling for main and top stories."""
        if main:
            # Extract main hero article details
            title = self.extract_text(article, "h1 a")
            link = self.extract_link(article, "h1 a", base_url=self.base_url)
            description = self.extract_text(article, "p.hero-article__teaser")
            date, content, image = self.extract_article_content(link)
        else:
            # Extract top story details
            title = self.extract_text(article, "h3 a")
            link = self.extract_link(article, "h3 a", base_url=self.base_url)
            description = None
            date, content, image = self.extract_article_content(link)

        # Return the Article object
        return Article(title, link, date, None, 'CyberSecurity Dive', description, image, content)
    
    def extract_article_content(self, article_url):
        """Fetch and extract content from the individual article page."""
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None, None, None

        soup = self.parse_html(page_content)

        # Extract the article date from the content page
        article_date = self.extract_article_date(soup)

        # Extract the full content of the article
        content = self.extract_content(soup, 'div.large.medium.article-body')
        
        # Extract the main image
        image_container = soup.find('div', class_='figure_content article-hero-img')
        image = self.extract_image(image_container, 'img', 'src')

        return article_date, content, image
    
    def extract_article_date(self, soup):
        # Locate the date element by its class name
        date_tag = soup.find('div', class_='date date-bottom-border')
        if not date_tag:
            date_tag = soup.find('div', class_='post-author-text-byline post-author-text-byline-date')
            if not date_tag:
                return 'No Date'

        # Extract and format the date string
        date_text = date_tag.get_text(strip=True).replace('Published ', '').replace('Updated ', '').strip()

        # Replace non-standard month abbreviations
        month_replacements = {
            "Sept.": "Sep."
        }

        for incorrect, correct in month_replacements.items():
            date_text = date_text.replace(incorrect, correct)

        date_text = re.sub(r'(\b\w+\.)', lambda match: match.group(1).replace('.', ''), date_text)
        return self.parse_date(date_text, date_format="%b %d, %Y")

    def extract_article_details(self, article):
        """Extract and return the details of a single article."""
        title = self.extract_text(article, 'h3.feed__title')
        link = self.extract_link(article, 'a', base_url=self.base_url)
        description = self.extract_text(article, 'p.feed__description')

        # Extract article content and date by visiting the article's page
        article_date, content, image = self.extract_article_content(link)
        if not self.is_article_valid(article_date):
            return None
        
        # Return the Article object
        return Article(title, link, article_date, None, 'CyberSecurity Dive', description, image, content)

    def scrape(self):
        """Main method to control the scraping process."""
        all_articles = []
        page_number = 1
        stop_scraping = False

        while not stop_scraping:
            # Construct the URL for the current page
            current_url = f"{self.base_url}?page={page_number}" if page_number > 1 else self.base_url

            # Fetch page content
            page_content = self.fetch_page_content(current_url)
            if not page_content:
                break

            # Parse the HTML content with BeautifulSoup
            soup = self.parse_html(page_content)

            # Extract articles from the current page
            articles = self.extract_articles(soup)
            if not articles:
                break
            elif page_number == 1:
                # Extract and add featured articles if it's the first page
                featured_articles = self.extract_featured_articles(soup)
                all_articles.extend(featured_articles)

            # Process and extract details from each article
            for article in articles:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    stop_scraping = True
                    break

            page_number += 1

        return all_articles

class BitdefenderScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the BitdefenderScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)
        self.load_more_clicks = 0  # Track the number of times 'Load More' has been clicked
        self.scraped_urls = set()  # Track the URLs of scraped articles

    def extract_articles(self, soup):
        """Extract article containers from the main page content."""
        return soup.find_all('div', class_='tw-mb-12 tw-flex-wrap tw-items-center sm:tw-flex')

    def extract_article_details(self, article):
        """Extract details from a single article and return an Article object."""
        # Extract the article link
        link = self.extract_link(article, 'a', base_url=self.base_url)
        if link in self.scraped_urls:
            return 'pass'
        self.scraped_urls.add(link) # Add the URL to the set of scraped URLs

        title = self.extract_text(article, 'h2', default='No title')

        article_date, content, image  = self.extract_article_content(link)
        if not article_date:
            return None  # Skip articles if we can't extract the date

        tags = [tag.text.strip() for tag in article.find_all('a', class_='tw-relative')][1:]

        return Article(title, link, article_date, tags, 'Bitdefender', content=content, image=image)

    def extract_article_content(self, article_url):
        """Visit the article page and extract content."""
        self.driver.get(article_url)
        soup = self.parse_html(self.driver.page_source)

        date_text = self.extract_text(soup, 'div.tw-pl-2 p')
        article_date = self.parse_date(date_text, date_format="%B %d, %Y")

        if not self.is_article_valid(article_date):
            return None, None, None
        
        content = self.extract_content(soup, 'div.content.tw-mb-12.tw-text-lg.tw-text-black')

        image = self.extract_image(soup, 'picture.tw-my-6.lazyLoad.isLoaded img', 'src')

        return article_date, content, image

    def click_load_more(self):
        """Click the 'Load More' button if available and increment the click counter."""
        try:
            # Wait for the "Load More" button to be clickable
            load_more_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Load more')]"))
            )
            # Click the "Load More" button
            load_more_button.click()
            time.sleep(3)  # Give time for the new content to load
            self.load_more_clicks += 1  # Keep track of how many times we've clicked it
            return True
        except Exception as e:
            print(f"Load More button not found or not clickable: {e}")
            return False

    def scrape(self):
        """Main method to control the scraping process using Selenium WebDriver."""
        all_articles = []  # Store all collected articles
        self.driver.get(self.base_url)
        time.sleep(3)  # Initial wait to ensure the page is fully loaded

        while True:
            # Parse the current page source using BeautifulSoup
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            article_elements = self.extract_articles(soup)

            # If no more articles are found, stop scraping
            if not article_elements:
                print("No more articles found.")
                break

            # Process each article element
            for article in article_elements:
                article_details = self.extract_article_details(article)
                if article_details == 'pass':
                    continue
                if article_details:
                    all_articles.append(article_details)
                else:
                    return all_articles
                
            self.last_url = all_articles[-1].url
            # Navigate back to the home page and click the "Load More" button the same number of times
            for _ in range(self.load_more_clicks + 1):
                self.driver.get(self.base_url)
                for _ in range(self.load_more_clicks):
                    if not self.click_load_more():
                        print("No more articles to load or 'Load More' button not clickable.")
                        return all_articles

            # Now, click the "Load More" button one more time to load new articles
            if not self.click_load_more():
                print("No more articles to load or 'Load More' button not clickable.")
                break

        return list(all_articles)

class SCADAfenceScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the SCADAfenceScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)
    
    def extract_articles(self, soup):
        """Extract articles from the main container in the soup object."""
        return soup.find_all('div', class_='postbox post-box post-item')

    def extract_article_details(self, article):
        """Extract and return the details of a single article."""
        date_tag = article.find('div', class_='date')
        article_date = self.parse_date(date_tag.get_text(strip=True), date_format="%B %d, %Y")

        if not self.is_article_valid(article_date):
            return None
        
        title = self.extract_text(article, 'div.title h4', default='No title')

        article_url = self.extract_link(article, 'a.img_link', base_url=self.base_url)

        image = self.extract_image(article, 'img.img-fluid', 'src')

        content = self.extract_article_content(article_url)

        return Article(title, article_url, article_date, None, 'SCADAfence', image=image, content=content)
    
    def extract_article_content(self, article_url):
        """Fetch and extract content from the individual article page."""
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None

        soup = self.parse_html(page_content)

        # Extract the main article content from paragraphs within the main content div
        content = self.extract_content(soup, 'div.section.post-body')

        return content

    def scrape(self):
        """Main method to scrape articles from the SCADAfence website."""
        all_articles = []  # To store all the articles

        # Fetch page content from the base URL (home page)
        page_content = self.fetch_page_content(self.base_url)
        if not page_content:
            return all_articles

        # Parse the HTML content using BeautifulSoup
        soup = self.parse_html(page_content)

        # Extract articles from the home page
        article_elements = self.extract_articles(soup)

        # Process each article and add to the list if valid
        for article in article_elements:
            article_details = self.extract_article_details(article)
            if article_details:
                all_articles.append(article_details)

        return all_articles

class ESecurityPlanetScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the scraper with base URL and minimum date, and set up categories."""
        super().__init__(base_url, min_date, driver)
        categories = ["cloud", "threats", "trends", "endpoint", "applications", "compliance", "cybersecurity", "products", "networks"]
        self.categories = [f"{base_url}{category}/" for category in categories]

    def extract_articles(self, soup):
        """Extract articles from the soup object."""
        # Locate the main article container using the given HTML structure
        return soup.find_all('li', class_='wp-block-post')

    def extract_article_details(self, article):
        """Extract and return the details of a single article."""
        title = self.extract_text(article, 'h4.wp-block-post-title', default='No title')

        article_url = self.extract_link(article, 'h4.wp-block-post-title a', base_url=self.base_url)

        # Extract the date
        article_date = self.extract_text(article, 'time')
        article_date = self.parse_date(article_date, date_format="%B %d, %Y")

        if not self.is_article_valid(article_date):
            return None

        description = self.extract_text(article, 'p.wp-block-post-excerpt__excerpt', default='No description')

        if article_url != '':
            content = self.extract_article_content(article_url)
        else:
            content = None

        return Article(title, article_url, article_date, None, 'ESecurityPlanet',description=description, content=content) 

    def extract_article_content(self, article_url):
        """Fetch and extract content from the individual article page."""
        if article_url == '':
            pass
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None

        soup = self.parse_html(page_content)

        content = self.extract_content(soup, 'div.entry-content.alignfull.wp-block-post-content.is-layout-flow.wp-block-post-content-is-layout-flow')

        return content  

    def get_next_page_url(self, current_url, page_number):
        """Construct the URL for the next page based on the current page number."""
        return f"{current_url}page/{page_number}/"

    def scrape_category(self, category_url):
        """Scrape a single category from the given URL."""
        all_articles = []
        page_number = 1

        while True:
            # Construct the URL for the current page
            current_url = self.get_next_page_url(category_url, page_number)

            # Fetch page content
            page_content = self.fetch_page_content(current_url)
            if not page_content:
                break

            # Parse the HTML content with BeautifulSoup
            soup = self.parse_html(page_content)
            articles = self.extract_articles(soup)

            # Process each article and check its date
            for article in articles:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    return all_articles

            # Move to the next page
            page_number += 1

        return all_articles

    def scrape(self):
        """Main method to scrape articles from multiple categories on ESecurityPlanet."""
        all_articles = []  # To hold articles from all categories

        # Scrape each category URL
        for category_url in self.categories:
            category_articles = self.scrape_category(category_url)
            all_articles.extend(category_articles)

        return all_articles

class SANSScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the SANSScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract articles from the soup object."""
        # Locate the main article containers using the class 'article-listing__item'
        return soup.find_all('li', class_='article-listing__item')

    def extract_article_details(self, article):
        """Extract and return the details of a single article."""
        # Extract the article title and link
        title = self.extract_text(article, 'div.title', default='No title')
        article_url = self.extract_link(article, 'a', base_url=self.base_url)

        # Extract and parse the date
        date = self.extract_text(article, 'div.date', default='No date')
        date = self.parse_date(date, date_format="%B %d, %Y")

        if not self.is_article_valid(date):
            return None

        description = self.extract_text(article, 'div.description.whitespace-break-spaces', default='No description')
        
        image = self.extract_image(article, 'img.img.img-square', 'src')

        content = self.extract_article_content(article_url)

        return Article(title, article_url, date, None, 'SANS', description, image, content)
    
    def extract_article_content(self, article_url):
        """Fetch and extract content from the individual article page."""
        self.driver.get(article_url)
        soup = self.parse_html(self.driver.page_source)

        # Extract the main article content from the 'div.article-content' element
        content = self.extract_content(soup, 'div.blog-content')

        return content

    def scrape(self):
        """Main method to control the scraping process using Selenium."""
        all_articles = []  # To store all articles from all pages
        page_number = 1  # Start from page 1

        while True:
            # Construct the URL for the current page using the page number
            current_url = f"{self.base_url}?page={page_number}"
            self.driver.get(current_url)
            time.sleep(3)  # Let the page load

            # Parse the page source using BeautifulSoup
            soup = self.parse_html(self.driver.page_source)
            articles = self.extract_articles(soup)

            # Process each article and check its date
            stop_scraping = False
            for article in articles:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    stop_scraping = True
                    break

            # Stop the loop if an old article is found or no articles are detected
            if stop_scraping or not articles:
                break

            # Increment the page number to move to the next page
            page_number += 1

        return all_articles

class DarkReadingScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the DarkReadingScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract articles from the soup object."""
        return soup.find_all('div', class_='ContentPreview')

    def extract_article_content(self, article_url):
        """Fetch and extract content and image from the individual article page."""
        page_content = self.fetch_page_content(article_url, True)
        if not page_content:
            return None

        soup = self.parse_html(page_content)
        content = self.extract_content(soup, 'div.ContentModule-Wrapper')

        return content

    def extract_article_details(self, article):
        """Extract details of a single article."""
        if article.find('div', class_='ContentPreview-Wrapper_variant_summary'):
            # Type 1: Featured article
            title = self.extract_text(article, 'a.ArticlePreview-Title', default='No Title')
            link = self.extract_link(article, 'a.ArticlePreview-Title', base_url=self.base_url)
            date_str = self.extract_text(article, 'span.ArticlePreview-Date', default='No Date')
            image = self.extract_image(article, 'div.ContentPreview-ImageContainer img', 'src')

        elif article.find('div', class_='ContentPreview-Wrapper_variant_normal'):
            # Type 2: Latests articles
            title = self.extract_text(article, 'a.ContentCard-Title', default='No Title')
            link = self.extract_link(article, 'a.ContentCard-Title', base_url=self.base_url)
            date_str = self.extract_text(article, 'span.ContentCard-Date', default='No Date')
            image = self.extract_image(article, 'div.ContentPreview-ImageContainer img', 'src')

        elif article.find('div', class_='ListPreview'):
            # Type 3: Article list
            title = self.extract_text(article, 'a.ListPreview-Title', default='No Title')
            link = self.extract_link(article, 'a.ListPreview-Title', base_url=self.base_url)
            date_str = self.extract_text(article, 'span.ListPreview-Date', default='No Date')
            image = self.extract_image(article, 'div.ListPreview-ImageWrapper img', 'src')

        else:
            # Unknown article type
            return "pass"

        date = self.parse_date(date_str, date_format="%b %d, %Y")
        if not self.is_article_valid(date):
            return None

        tags = [self.extract_text(article, 'a.Keyword', default='No Tag')]
        content = self.extract_article_content(link)

        return Article(title, link, date, tags, 'Dark Reading', image=image, content=content)

    def scrape(self):
        """Main method to control the scraping process."""
        all_articles = []
        page_number = 1

        while True:
            current_url = f"{self.base_url}?page={page_number}" if page_number > 1 else self.base_url
            page_content = self.fetch_page_content(current_url)
            if not page_content:
                break

            soup = self.parse_html(page_content)
            articles = self.extract_articles(soup)

            if not articles:
                break

            for article in articles:
                article_details = self.extract_article_details(article)
                if article_details == "pass":
                    continue
                elif article_details:
                    all_articles.append(article_details)
                else:
                    return all_articles

            page_number += 1

        return all_articles

class TenableScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract article containers from the page."""
        return soup.find_all('div', class_='blog-item')

    def extract_article_details(self, article):
        """Extract details of a single article."""
        title = self.extract_text(article, 'h2 a', default='No title')
        link = self.extract_link(article, 'h2 a', base_url=self.base_url)

        date_str = self.extract_text(article, 'em', default='No Date')
        date = self.parse_date(date_str, "%B %d, %Y")

        if not self.is_article_valid(date):
            return None

        image_style = article.find('div', class_='blog-item__image')
        image = re.search(r'url\((.*?)\)', image_style['style']).group(1) if image_style else None
        image = urljoin(self.base_url, image) if image else None

        content = self.extract_article_content(link)

        return Article(title, link, date, None, 'Tenable', None, image, content)

    def extract_article_content(self, article_url):
        """Fetch and extract content and full image from the article page."""
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None

        soup = self.parse_html(page_content)
        content = self.extract_content(soup, 'div.blog__body')
        return content

    def scrape(self):
        """Main scraping function."""
        all_articles = []

        page_content = self.fetch_page_content(self.base_url)
        if not page_content:
            return all_articles
        
        soup = self.parse_html(page_content)
        articles = self.extract_articles(soup)

        for article in articles:
            if str(article) == '<div class="blog-item featured-content featured-post"></div>':
                continue
            article_details = self.extract_article_details(article)
            if article_details:
                all_articles.append(article_details)
            else:
                return all_articles
        
        current_page = -1
        while True:
            current_page += 1
            page_url = self.base_url + f'/all?page={current_page}'
            page_content = self.fetch_page_content(page_url)
            if not page_content:
                return all_articles
            
            soup = self.parse_html(page_content)
            articles = self.extract_articles(soup)

            for article in articles:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    return all_articles

class Presale1Scraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the Presale1Scraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)
        self.load_more_clicks = 0  # Track the number of times 'Load More' has been clicked
        self.scraped_urls = set()  # Track the URLs of scraped articles

    def extract_articles(self, soup):
        """Extract article containers from the main page content."""
        articles = soup.find('div', attrs={'data-ux': 'Grid', 'data-aid': 'RSS_FEEDS_RENDERED'})
        return articles.find_all('div', attrs={'data-ux': 'GridCell'})

    def extract_article_details(self, article):
        """Extract details from a single article and return an Article object."""
        date_element = article.find('p', attrs={'data-aid': 'RSS_FEED_POST_DATE_RENDERED'})
        date_str = date_element.get_text(strip=True) if date_element else 'No Date'
        date = self.parse_date(date_str, "%d %B %Y")
        if not self.is_article_valid(date):
            return None 
        
        # Extract the article link
        link = self.extract_link(article, 'a', base_url=self.base_url)
        if link in self.scraped_urls:
            return 'pass'
        self.scraped_urls.add(link)  # Add the URL to the set of scraped URLs

        title = self.extract_text(article, 'h4', default='No title')

        content, image = self.extract_article_content(link)

        return Article(title, link, date, None, 'Presale1', None, image, content)

    def extract_article_content(self, article_url):
        """Visit the article page and extract content."""
        self.driver.get(article_url)
        soup = self.parse_html(self.driver.page_source)

        content = self.extract_content(soup, 'div[data-ux="BlogContent"]')

        content_section = soup.select_one('div[data-ux="BlogContent"]')
        image = self.extract_image(content_section, 'img', 'src')
        image = 'https:' + image if image else None

        return content, image

    def click_load_more(self):
        """Click the 'Load More' button if available and increment the click counter."""
        try:
            # Wait for the "Load More" button to be clickable
            load_more_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Show More')]"))
            )
            # Click the "Load More" button
            load_more_button.click()
            time.sleep(3)  # Give time for the new content to load
            self.load_more_clicks += 1  # Keep track of how many times we've clicked it
            return True
        except Exception as e:
            print(f"Load More button not found or not clickable: {e}")
            return False

    def scrape(self):
        """Main method to control the scraping process using Selenium WebDriver."""
        all_articles = []  # Store all collected articles
        self.driver.get(self.base_url)
        time.sleep(3)  # Initial wait to ensure the page is fully loaded

        while True:
            # Parse the current page source using BeautifulSoup
            soup = self.parse_html(self.driver.page_source)
            article_elements = self.extract_articles(soup)

            # If no more articles are found, stop scraping
            if not article_elements:
                print("No more articles found.")
                break

            # Process each article element
            for article in article_elements:
                article_details = self.extract_article_details(article)
                if article_details == 'pass':
                    continue
                if article_details:
                    all_articles.append(article_details)
                else:
                    return all_articles

            # Navigate back to the home page and click the "Load More" button the same number of times
            for _ in range(self.load_more_clicks + 1):
                self.driver.get(self.base_url)
                for _ in range(self.load_more_clicks):
                    if not self.click_load_more():
                        print("No more articles to load or 'Load More' button not clickable.")
                        return all_articles

            # Now, click the "Load More" button one more time to load new articles
            if not self.click_load_more():
                print("No more articles to load or 'Load More' button not clickable.")
                break

        return list(all_articles)

class SectrioScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract articles from the main articles container."""
        articles_container = soup.find('div', class_='elementor-posts-container')
        return articles_container.find_all('article', class_='elementor-post') if articles_container else []

    def extract_article_details(self, article):
        """Extract details of a single article."""
        # Extract title and URL
        title = self.extract_text(article, 'h3.elementor-post__title a', default='No Title')
        link = self.extract_link(article, 'h3.elementor-post__title a', base_url=self.base_url)

        # Extract description
        description = self.extract_text(article, 'div.elementor-post__excerpt p', default='No Description')

        article_date, content, image = self.extract_article_content(link)

        # Validate date
        if not self.is_article_valid(article_date):
            return None

        return Article(title, link, article_date, None, 'Sectrio', description, image, content)

    def extract_article_content(self, article_url):
        """Fetch and extract content from the individual article page."""
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None

        soup = self.parse_html(page_content)
        meta_tag = soup.find('meta', attrs={'property': 'article:published_time'})
        # Extract the content attribute if the tag exists
        if meta_tag and 'content' in meta_tag.attrs:
            date_str = meta_tag['content']
            # Parse the date string
            parsed_date = self.parse_date(date_str, date_format="%Y-%m-%dT%H:%M:%S%z")
        else:
            print("Published time not found.")

        content = self.extract_content(soup, 'div.bialty-container')
        image = self.extract_image(soup, 'figure.wp-block-image.size-large img', 'data-src')

        return parsed_date, content, image

    def scrape(self):
        """Main method to control the scraping process."""
        all_articles = []
        current_url = self.base_url
        current_page = 1

        while True:
            page_content = self.fetch_page_content(current_url)

            if not page_content:
                return all_articles

            soup = self.parse_html(page_content)
            articles = self.extract_articles(soup)

            if not articles:
                return all_articles

            for article in articles:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    return all_articles
            
            current_page += 1
            current_url = f"{self.base_url}{current_page}/"

class Intel471Scraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        super().__init__(base_url, min_date, driver)

    def extract_articles(self, soup):
        """Extract articles from the main container."""
        return soup.find_all('a', target='_self')

    def extract_article_details(self, article):
        """Extract details of a single article."""
        # Extract title
        title = self.extract_text(article, 'h3.h4.mb-3', default='No Title')

        # Extract link
        link = self.extract_link(article, '', base_url=self.base_url)

        # Extract description
        description = self.extract_text(article, 'p', default='No Description')

        # Extract date
        date_text = self.extract_text(article, 'div.body-overline-small.text-neutral-500', default='')
        date_str = date_text.split('//')[-1].strip()
        tags = date_text.split('//')[:-1]
        article_date = self.parse_date(date_str, date_format="%b %d, %Y")

        # Validate date
        if not self.is_article_valid(article_date):
            return None

        # Extract image
        image = None

        content = self.extract_article_content(link)
        
        return Article(title, link, article_date, tags, 'Intel471', description, image, content)

    def extract_article_content(self, article_url):
        """Fetch and extract content from the individual article page."""
        page_content = self.fetch_page_content(article_url)
        if not page_content:
            return None

        soup = self.parse_html(page_content)
        content = self.extract_content(soup, "div.rich-text")
        return content

    def get_next_page_url(self, soup):
        """Find and return the URL for the next page."""
        next_page = soup.find('a', class_='block hover:text-red-700 text-base no-underline cursor-pointer', attrs={'data-hx-push-url': True})
        return self.extract_link(next_page, '', base_url=self.base_url) if next_page else None

    def scrape(self):
        """Main method to control the scraping process."""
        all_articles = []
        next_page_url = self.base_url
        current_page = 1

        while next_page_url:
            page_content = self.fetch_page_content(next_page_url)
            if not page_content:
                break

            soup = self.parse_html(page_content)
            article_div = soup.find('div', id='listing-results')
            articles = self.extract_articles(article_div)

            for article in articles:
                article_details = self.extract_article_details(article)
                if article_details:
                    all_articles.append(article_details)
                else:
                    return all_articles

            current_page += 1
            next_page_url = f"{self.base_url}?page={current_page}"

        return all_articles

class IndustrialCyberScraper(BaseScraper):
    def __init__(self, base_url, min_date, driver=None):
        """Initialize the IndustrialCyberScraper with base URL and minimum date."""
        super().__init__(base_url, min_date, driver)
        self.categories = ["news", "features"]
    
    def extract_articles(self, soup):
        """Implement abstract method to satisfy BaseScraper requirements."""
        return []
    
    def extract_article_content(self, link):
        """Extract the date and content of an article."""
        self.driver.get(link)
        soup = self.parse_html(self.driver.page_source)

        # Extract the article's publication date
        date_tag = soup.find('div', class_='single-post-date')
        date = self.parse_date(date_tag.get_text(strip=True), date_format="%B %d, %Y")

        # Extract the main content of the article
        content = self.extract_content(soup, 'div.single-content')

        return date, content

    def extract_articles_from_breaking_news(self, soup):
        """Extract articles from the 'breaking-news' section."""
        articles = []

        soup = self.parse_html(self.driver.page_source)
        breaking_news_section = soup.find('div', class_='breaking-news')
        if not breaking_news_section:
            print("No 'breaking-news' section found.")
            return articles

        # Locate the article previews
        previews = breaking_news_section.find_all('a')
        if not previews:
            print("No previews found in 'breaking-news'.")
            return articles

        for preview in previews:
            try:
                # Extract article details
                link = preview['href']
                title = self.extract_text(preview, 'div.preview-title', default='No Title')
                description = None
                image = self.extract_image(preview, 'div.preview-thumb img', 'src')
                date, content = self.extract_article_content(link)

                if self.is_article_valid(date):
                    articles.append(Article(title, link, date, None, 'IndustrialCyber', description, image, content))
                else:
                    break
                
            except Exception as e:
                print(f"Error extracting article: {e}")
        return articles

    def extract_articles_from_archive_posts(self, soup):
        """Extract articles from the 'archive-posts' section."""
        articles = []
        archive_posts_section = soup.find('div', class_='archive-posts')
        if not archive_posts_section:
            print("No 'archive-posts' section found.")
            return articles

        previews = archive_posts_section.find_all('div', class_='related-preview')
        for preview in previews:
            try:
                # Extract article details
                link = self.extract_link(preview, 'a', base_url=self.base_url)
                title = self.extract_text(preview, 'div.related-preview-title', default='No Title')
                description = None
                image = self.extract_image(preview, 'div.related-preview-thumb img', 'src')
                date, content = self.extract_article_content(link)

                if self.is_article_valid(date):
                    articles.append(Article(title, link, date, None, 'IndustrialCyber', description, image, content))
                else:
                    articles.append(None)
                    break
                
            except Exception as e:
                print(f"Error extracting article from 'archive-posts': {e}")
        return articles

    def scrape_category(self, category):
        """Scrape articles from a specific category."""
        all_articles = []
        category_url = f"{self.base_url}{category}/"
        next_count = 0

        self.driver.get(category_url)
        soup = self.parse_html(self.driver.page_source)

        # Extract articles from 'breaking-news'
        breaking_news_articles = self.extract_articles_from_breaking_news(soup)
        all_articles.extend(breaking_news_articles)

        self.driver.get(category_url)

        while True:
            soup = self.parse_html(self.driver.page_source)
            
            # Extract articles from 'archive-posts'
            archive_posts_articles = self.extract_articles_from_archive_posts(soup)
            all_articles.extend(archive_posts_articles)

            if all_articles[-1] is None:
                all_articles.pop()
                return all_articles

            # Try to click the "Next" button for pagination
            self.driver.get(category_url)
            next_count += 1
            for _ in range(next_count):
                if not self.click_next_page():
                    return all_articles

    def click_next_page(self):
        """Simulate a click on the 'Next' button to load more articles."""
        try:
            # Wait for the button to appear on the page
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'a.next.page-numbers'))
            )
            
            # Locate the next button
            next_button = self.driver.find_element(By.CSS_SELECTOR, 'a.next.page-numbers')

            # Ensure the button is displayed and enabled
            if next_button.is_displayed() and next_button.is_enabled():
                self.driver.execute_script("arguments[0].click();", next_button)
                return True
            else:
                print("Next button is not clickable.")
                return False
        except Exception as e:
            print(f"No 'Next' button found or clickable: {e}")
            return False

    def scrape(self):
        """Main method to scrape articles from all categories."""
        all_articles = []
        for category in self.categories:
            category_articles = self.scrape_category(category)
            all_articles.extend(category_articles)
        return all_articles

# Main functions --------------------------------------------------
def load_configurations(config_file=CONFIGS_FILE):
    """Load scraper configurations from a CSV file, dynamically linking to the scraper classes."""
    configs = []
    with open(config_file, mode='r', newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            # Dynamically resolve the scraper class using the class name from CSV
            scraper_class_name = row['scraper_class']
            scraper_class = globals().get(scraper_class_name)  # Attempt to find the class in global scope

            if scraper_class is None:
                print(f"Scraper class '{scraper_class_name}' not found. Skipping this configuration.")
                continue  # Skip configurations with unresolved scraper classes

            configs.append({
                'name': row['name'],
                'scrape_type': row['scrape_type'],
                'url': row['url'],
                'scraper_class': scraper_class  # Store the actual class reference
            })
    return configs

def clear_screen():
    """Clear the terminal screen."""
    if os.name == 'nt':  # For Windows
        os.system('cls')
    else:  # For macOS and Linux
        os.system('clear')

def get_user_input(configs):
    """Get user input for selecting websites and minimum date."""
    # Show the list of available websites
    available_sites = [config['name'] for config in configs]
    print("Available websites to scrape:")

    # Display websites in a numbered list with a two-column layout for better readability
    for index, site in enumerate(available_sites, start=1):
        # Create two columns of website names
        if index % 2 == 1:
            # Print site on the left column
            print(f"{index:2}. {site:<30}", end="")
        else:
            # Print site on the right column
            print(f"{index:2}. {site:<30}")
            
    # Handle the case when the number of websites is odd
    if len(available_sites) % 2 == 1:
        print()  # Ensure the final line ends correctly

    # Get user input for which websites to scrape
    selected_input = input("\nEnter the number(s) of website(s) to scrape (comma-separated or '0' for all websites): ").strip()
    selected_sites = selected_input.split(',')

    # Determine which configurations to select based on user input
    if '0' in selected_sites:
        selected_configs = configs  # Select all websites
    else:
        try:
            # Convert input to a set of indices and validate against available sites
            selected_indices = {int(site.strip()) for site in selected_sites if site.strip().isdigit()}
            selected_configs = [configs[index - 1] for index in selected_indices if 1 <= index <= len(configs)]
        except ValueError:
            print("Invalid input. Please enter numbers only.")
            return None, None

    if not selected_configs:
        print("No valid websites selected. Exiting...")
        return None, None

    # Get user input for minimum date
    min_date_str = input("Enter the minimum date to consider for articles (format: DD-MM-YY): ").strip()
    try:
        min_date = datetime.strptime(min_date_str, "%d-%m-%y").date()
    except ValueError:
        print(f"Invalid date format. Using default date: '{MIN_DATE.strftime('%d-%m-%Y')}'.")
        min_date = MIN_DATE

    return selected_configs, min_date

def scrape_selected_sites(configs, min_date):
    """Scrape articles from the selected websites using class-based scrapers."""
    print("\nScraping in progress...")

    all_articles = set()
    driver = BaseScraper.create_webdriver(None)
    errors = []
    
    for config in tqdm(configs, desc="Scraping websites", unit="site"):
        try:
            scraper = config['scraper_class'](config['url'], min_date, driver)
            articles = set(scraper.scrape())
            all_articles.update(articles)
            print(f"\nSuccessfully scraped {len(articles)} articles from {config['name']}")
            
        except Exception as e:
            print(f"\nError scraping {config['name']}: {str(e)}")
            errors.append((config['name'], str(e)))
            continue  # Continue with next website even if current one fails
    
    print("\nScraping completed.")

    if errors:
        print("\nErrors occurred while scraping the following websites:")
        for site, error in errors:
            print(f"{site}: {error}")

    driver.quit()
    return all_articles

def save_results(articles, output_file=OUTPUT_FILE):
    """Save the scraped articles to an Excel file with clickable links."""
    output_dir = os.path.dirname(output_file)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)  # Create the directory if it does not exist

    # Convert each Article object into a dictionary
    articles_data = [vars(article) for article in articles]
    articles_df = pd.DataFrame(articles_data)

    # Use ExcelWriter to save with specific formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        articles_df.to_excel(writer, index=False, sheet_name='Articles')

        # Access the workbook and worksheet for hyperlink insertion
        worksheet = writer.sheets['Articles']

        # Make URLs clickable in the "url" column
        for row_idx, url in enumerate(articles_df['url'], start=2):  # Start at row 2 to skip header
            worksheet.cell(row=row_idx, column=2).hyperlink = url
            worksheet.cell(row=row_idx, column=2).style = "Hyperlink"
        
        # Make URLs clickable in the "image" column 
        for row_idx, url in enumerate(articles_df['image'], start=2):  # Start at row 2 to skip header
            worksheet.cell(row=row_idx, column=7).hyperlink = url
            worksheet.cell(row=row_idx, column=7).style = "Hyperlink"

        # Apply date format to the "date" column
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YY")
        for row_idx in range(2, len(articles_df) + 2):  # Adjust for header row
            worksheet.cell(row=row_idx, column=3).style = date_style  # Assuming the date is in column 3

    print(f"\n{len(articles)} articles saved to {os.path.abspath(output_file)}.")

def open_file(file_path):
    """Open the given file based on the user's operating system."""
    try:
        if platform.system() == 'Windows':
            os.startfile(file_path)  # For Windows
        elif platform.system() == 'Darwin':
            subprocess.call(['open', file_path])  # For macOS
        elif platform.system() == 'Linux':
            subprocess.call(['xdg-open', file_path])  # For Linux
        else:
            print(f"Unsupported OS: {platform.system()}. Cannot open the file automatically.")
    except Exception as e:
        print(f"Failed to open file {file_path}: {e}")

def main():
    """Main execution function for the script."""
    # Load configurations and get user input
    configs = load_configurations()
    selected_configs, min_date = get_user_input(configs)

    if not selected_configs:
        return  # Exit if no valid configurations were selected

    clear_screen()

    # Start scraping and measure the total runtime
    start_time = time.time()

    # Scrape the selected websites
    all_articles = scrape_selected_sites(selected_configs, min_date)

    # Save the results if there are any articles scraped
    if all_articles:
        save_results(all_articles, OUTPUT_FILE)

        end_time = time.time()
        total_runtime = end_time - start_time
        print(f"\nScript completed in {total_runtime:.2f} seconds.")

        # Ask the user if they want to open the CSV file
        open_csv = input("\nDo you want to open the file now? (y/n): ").strip().lower()
        if open_csv in ['y', 'yes']:
            open_file(OUTPUT_FILE)
    else:
        print("No articles to save. Exiting...")

if __name__ == "__main__":
    main()



